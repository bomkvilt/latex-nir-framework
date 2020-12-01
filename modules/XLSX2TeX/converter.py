from openpyxl import load_workbook


class XLSX2TeXConverter:
    def __init__(self):
        pass

    def convert(self, sheet_path:str) -> str:
        wb = load_workbook(sheet_path, data_only=True)
        ws = wb.active
        mg = ws.merged_cells.ranges

        width = 0
        for cell in ws[1]:
            if cell.value != None:
                width += 1
            else: break
        
        text  = r'\begin{tabu}{|' + 'X[c,m]|' * width + '} '
        text += r'\hline' + '\n'
        for row in ws.iter_rows():
            line   = []
            bEmpty = True
            for cell in row[:width]:
                value   = cell.value
                cformat = cell.number_format
                if cformat.startswith("0"):
                    parts   = cformat.split(".")
                    count   = 0 if len(parts) == 1 else len(parts[1])
                    pattern = "{:." + str(count) + "f}"
                    value   = pattern.format(value)
                else:
                    value = str(value)
                line.append(self.sheeld(value))
                bEmpty &= cell.value == None
            text += '\t'
            text += ' & '.join(line)
            text += r'\\ \hline'
            text += '\n'
        text += r'\end{tabu}'
        return text

    def sheeld(self, val:str) -> str:
        for f, t in {
            '%': r'\%',
            '^': r'\^',
        }.items():
            val = val.replace(f, t)
        return val
